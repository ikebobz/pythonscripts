from playwright.sync_api import sync_playwright
import pandas as pd
import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def login_to_website(url, username, password):
    with sync_playwright() as p:
        # Launch the browser in a non-headless mode to see the actions
        # For production or testing, you can set headless=True
        browser = p.chromium.launch(headless=False, slow_mo=500)
        page = browser.new_page()
        today = datetime.datetime.now()
        today_str = f"{today.strftime('%b')} {today.day}, {today.year}"
        target_date = (datetime.datetime.now() - datetime.timedelta(days=2)).strftime("%b %d, %Y")
        selection = [
            "LagosGeneralHospital",
            "RandleGeneralHospital",
            "RCCG",
            "ApapaGeneralHospital",
            "AhamadiyyaMuslimHospital",
            "SacredHeartClinic",
            "IjoraPrimaryHealthCentre",
            "Coker",
            "Baruwa",
            "LagosIslandMaternityHospital",
            "HarveyRoadGeneralHospital",
            "Massey",
            "Onikan",
            "Olojowon",
            "Simpson",
            "Akerele",
            "Ejire",
            "SuraPrimaryHealthCentre",
            "IgaIdugaranPrimaryHealthCentre",
            "NigerianInstituteofMedicalResearch"        
            ]
        


        # Navigate to the login page
        page.goto(url)

        # Locate the username and password fields and fill them
        # You need to inspect the website's HTML to find the correct selectors (e.g., #username, .password-field)
        page.locator("#Email").fill(username)
        page.locator("#Password").fill(password)
        
        # Locate and click the login button
        # This can be a button with a specific text or a CSS selector
        page.locator("#loginButton").click()

        # Wait for the next page to load after a successful login
        page.wait_for_url("https://ndr.nascp.gov.ng/uploads/partner-uploads")
        
        # Use a Playwright locator to find the table by its ID
        table = page.locator("#uploadDataTable")
        
        # Wait for the table to be visible on the page
        table.wait_for(state="visible")

        # Get the header row and its cells (<th>)
        header_row = table.locator("thead tr").nth(1).locator("th")
        headers = ['IP','Batch','Status','Total','Fails','Passes','Pending','Logs']
        
        #wait for at least one row of data to be present
        table.locator("tbody tr").first.wait_for(state="visible", timeout=5000)
        
        # Get all the data rows (<tr>) and their cells (<td>)
        data_rows = table.locator("tbody tr")
        
        table_data = []
        # ...existing code...
        page_count = 0
        max_pages = 50 # Set a limit to avoid infinite loops in case of unexpected behavior
        
        while True:
            # Wait for at least one row of data to be present
            stop_due_to_date = False
            table.locator("tbody tr").first.wait_for(state="visible", timeout=5000)
            data_rows = table.locator("tbody tr")
            for row in data_rows.all():
                row_cells = row.locator("td").all_text_contents()
                if not (row_cells and today_str in row_cells[0]):
                    stop_due_to_date = True
                    break
                if any(sel in row_cells[1] for sel in selection ):
                    table_data.append(row_cells)
            if stop_due_to_date:
                break
            page_count += 1
            if page_count >= max_pages:
                #print("Reached maximum page limit.")
                break    

            # Check if the "Next" button is enabled/visible
            next_button = page.locator("#uploadDataTable_next").first
            if next_button.is_disabled() or not next_button.is_visible():
                break  # No more pages

            # Click the "Next" button and wait for the table to update
            next_button.click()
            page.wait_for_timeout(10000)  # Adjust timeout as needed

        
        print(f"Extracted {len(table_data)} rows of data.")
        
        # Create a pandas DataFrame from the extracted data
        df = pd.DataFrame(table_data, columns=headers)

    # Save the DataFrame to an Excel file
        try:
          #df.to_excel("uploads.xlsx", index=False)
          generate_daily_tracker(df)
          #print(f"Successfully downloaded table ")
        except Exception as e:
          print(f"Failed to save Excel file: {e}")
        
    

        # Keep the browser open for a few seconds to see the result
        page.wait_for_timeout(3000)

        browser.close()
# ...existing code...

def generate_daily_tracker(df):
    """
    Extracts the string between the second and third underscore in the second column of each row,
    counts occurrences, and saves the result to 'daily_tracker.xlsx'.
    """
    # Extract the target substring from the second column
    def extract_between_underscores(s):
        parts = s.split('_')
        if len(parts) > 5:
            # Join parts between the third (index 3) and fifth (index 5) underscore
            return '_'.join(parts[3:5])
        return None

    # Apply extraction to the second column
    extracted = df.iloc[:, 1].apply(extract_between_underscores)
    # Count occurrences
    counts = extracted.value_counts().reset_index()
    counts.columns = ['UniqueValue', 'Occurrences']
    outputdf = extract_facility_and_services(counts)
    pivoteddf = pivotDataFrame(outputdf)

    # Save to Excel
    pivoteddf.to_excel("daily_tracker.xlsx", index=False)
    
    # Apply cell coloring for yes/no values
    color_cells_by_value("daily_tracker.xlsx")
    print("Saved daily_tracker.xlsx")

# ...existing code...
def extract_facility_and_services(df):
    """
    Processes the first column of the input DataFrame.
    Output DataFrame columns: Facility, treatment, hts, bio.
    For each row:
      - Facility: word before the first underscore in the first cell.
      - treatment/hts/bio: 'yes' if the word after the first underscore contains 'treatment', 'hts', or 'bio', else 'no'.
    """
    facilities = []
    treatments = []
    hts_list = []
    bios = []

    for val in df.iloc[:, 0]:
        if isinstance(val, str) and '_' in val:
            parts = val.split('_', 1)
            facility = parts[0]
            after_underscore = parts[1].lower()
            facilities.append(facility)
            treatments.append('yes' if 'treatment' in after_underscore else 'no')
            hts_list.append('yes' if 'hts' in after_underscore else 'no')
            bios.append('yes' if 'bio' in after_underscore else 'no')
        else:
            facilities.append(None)
            treatments.append('no')
            hts_list.append('no')
            bios.append('no')

    result_df = pd.DataFrame({
        'Facility': facilities,
        'treatment': treatments,
        'hts': hts_list,
        'bio': bios
    })
    return result_df

def pivotDataFrame(df):
    
    def collapse_services(series):
     return "yes" if "yes" in series.values else "no"

    collapsed = df.groupby("Facility").agg({
    "treatment": collapse_services,
    "hts": collapse_services,
    "bio": collapse_services
     }).reset_index()

    return collapsed

def color_cells_by_value(excel_file):
    """
    Colors cells in the Excel file green if they contain 'yes' and red if they contain 'no'.
    """
    # Define colors
    green_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")  # Green
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")    # Red
    
    # Load the workbook
    wb = load_workbook(excel_file)
    ws = wb.active
    
    # Iterate through all cells and apply coloring based on content
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.value == "yes":
                cell.fill = green_fill
            elif cell.value == "no":
                cell.fill = red_fill
    
    # Save the workbook
    wb.save(excel_file)
    
    
if __name__ == "__main__":
    # Replace these with the actual website details
    LOGIN_URL = "https://ndr.nascp.gov.ng/uploads/partner-uploads"
    USERNAME = "aomoaka@ecews.org"
    PASSWORD = "Pass1word#"
    
    login_to_website(LOGIN_URL, USERNAME, PASSWORD)
    #df = pd.read_excel("uploads.xlsx")
    #generate_daily_tracker(df)